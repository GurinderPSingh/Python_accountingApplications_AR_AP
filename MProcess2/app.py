import os
import subprocess
from flask import Flask, request, render_template, jsonify
import openpyxl
from Mpworkingdata import load_workbook_and_sheet, get_unique_entries_in_column, move_rows_based_on_selection

app = Flask(__name__)
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
SCRIPTS_FOLDER = os.getcwd()  # Assuming the scripts are in the current working directory
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload_file', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No file part in the request", 400

    file = request.files['file']
    if file.filename == '':
        return "No file selected", 400

    if file:
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet_names = wb.sheetnames
            return jsonify({"file_path": file_path, "sheet_names": sheet_names})
        except Exception as e:
            return f"Error reading file: {e}", 500

@app.route('/get_unique_values', methods=['POST'])
def get_unique_values():
    try:
        data = request.json
        file_path = data.get('file_path')
        sheet_name = data.get('sheet_name')

        if not os.path.isfile(file_path):
            return "File does not exist", 400

        wb, sheet = load_workbook_and_sheet(file_path, sheet_name)
        unique_values = get_unique_entries_in_column(sheet)

        return jsonify({"unique_values": unique_values})
    except Exception as e:
        return f"Error reading file or fetching unique values: {e}", 500

@app.route('/process_unique_values', methods=['POST'])
def process_unique_values():
    try:
        data = request.json
        file_path = data.get('file_path')
        sheet_name = data.get('sheet_name')
        selected_values = data.get('selected_values')

        if not os.path.isfile(file_path):
            return "File does not exist", 400

        wb, sheet = load_workbook_and_sheet(file_path, sheet_name)
        move_rows_based_on_selection(sheet, selected_values)

        # Save the updated workbook
        wb.save(file_path)

        return "Rows processed successfully"
    except Exception as e:
        return f"Error processing unique values: {e}", 500

@app.route('/run_script', methods=['POST'])
def run_script():
    data = request.get_json()
    file_path = data.get('file_path')
    sheet_name = data.get('sheet_name')
    script_name = data.get('script_name')

    if not file_path or not os.path.isfile(file_path):
        return "File does not exist or is not specified", 400

    if not sheet_name:
        return "Sheet name not provided", 400

    if not script_name:
        return "Script name not provided", 400

    # Get the full path of the script
    script_path = os.path.join(SCRIPTS_FOLDER, script_name)
    if not os.path.isfile(script_path):
        return f"Script {script_name} does not exist in the expected location.", 400

    try:
        print(f"Running {script_name} with file: {file_path}, sheet: {sheet_name}")
        subprocess.run(["python", script_path, file_path, sheet_name], check=True)
        return f"Script {script_name} executed successfully."
    except subprocess.CalledProcessError as e:
        return f"Error running script: {e.stderr}", 500
    except Exception as e:
        return f"An unexpected error occurred: {e}", 500

@app.route('/run_perk_script', methods=['POST'])
def run_perk_script():
    if 'perk-source' not in request.files or 'perk-file' not in request.files:
        return "Perk source file or perk file is missing.", 400

    source_file = request.files['perk-source']
    perk_file = request.files['perk-file']
    destination_folder = request.form.get('destination-folder')
    destination_name = request.form.get('destination-file')

    if not source_file or not source_file.filename:
        return "Perk source file is missing or invalid.", 400

    if not perk_file or not perk_file.filename:
        return "Perk file is missing or invalid.", 400

    if not destination_folder or not destination_name:
        return "Destination folder or file name is missing.", 400

    try:
        source_path = os.path.join(UPLOAD_FOLDER, source_file.filename)
        perk_path = os.path.join(UPLOAD_FOLDER, perk_file.filename)
        destination_path = os.path.join(destination_folder, destination_name)

        # Save the uploaded files
        source_file.save(source_path)
        perk_file.save(perk_path)

        # Run the Perk script
        subprocess.run(
            ["python", os.path.join(SCRIPTS_FOLDER, "Perk.py"),
             source_path, perk_path, destination_path, "WorkingSheet"],
            check=True
        )

        return f"Script executed successfully. Output saved as: {destination_path}"
    except subprocess.CalledProcessError as e:
        return f"Error running Perk script: {e.stderr}", 500
    except Exception as e:
        return f"An unexpected error occurred: {e}", 500


if __name__ == '__main__':
    # app.run(debug=True)
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))