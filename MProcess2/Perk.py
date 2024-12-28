import openpyxl
import shutil
import os


def load_workbook(file_path):
    """
    Load an Excel workbook.
    """
    return openpyxl.load_workbook(file_path)


def add_perk_dates_to_working_sheet(source_file_path, perk_file_path, new_file_path, working_sheet_name="WorkingSheet"):
    """
    Add dates from 'Perk' workbook (Col J, ITS793 sheet) to entries in Column W of the WorkingSheet
    if they match entries in Column A of the WorkingSheet with Column D of Perk workbook.

    Args:
        source_file_path (str): Path to the original source_modified workbook.
        perk_file_path (str): Path to the Perk workbook.
        new_file_path (str): Path to save the new workbook as Perk_source_modified.
        working_sheet_name (str): Name of the sheet in the WorkingSheet workbook to process.
    """
    print(f"Creating a copy of '{source_file_path}' as '{new_file_path}'...")

    # Ensure destination folder exists
    destination_folder = os.path.dirname(new_file_path)
    if not os.path.exists(destination_folder):
        raise ValueError(f"Destination folder does not exist: {destination_folder}")

    shutil.copyfile(source_file_path, new_file_path)

    print("Loading 'WorkingSheet' workbook...")
    working_wb = load_workbook(new_file_path)
    if working_sheet_name not in working_wb.sheetnames:
        raise ValueError(f"The sheet '{working_sheet_name}' does not exist in the working workbook.")
    working_sheet = working_wb[working_sheet_name]

    print("Loading 'Perk' workbook...")
    perk_wb = load_workbook(perk_file_path)
    if "ITS793" not in perk_wb.sheetnames:
        raise ValueError("The 'Perk' workbook does not contain a sheet named 'ITS793'.")
    perk_sheet = perk_wb["ITS793"]

    # Extract all entries from Column D and corresponding dates from Column J in 'Perk' workbook into a dictionary
    print("Extracting entries from 'Perk' workbook...")
    perk_data = {}
    for row in perk_sheet.iter_rows(min_row=2, max_row=perk_sheet.max_row, min_col=4, max_col=10):
        key_value = row[0].value  # Column D (key)
        date_value = row[6].value  # Column J (date)
        if key_value is not None and date_value is not None:
            perk_data[str(key_value).strip()] = date_value

    # Add a new column header for "Perk_tag" in Column W if not already present
    print("Adding 'Perk_tag' column to 'WorkingSheet'...")
    col_w_index = 23  # Column W is the 23rd column
    if working_sheet.cell(row=1, column=col_w_index).value != "Perk_tag":
        working_sheet.cell(row=1, column=col_w_index, value="Perk_tag")

    # Iterate over each row in 'WorkingSheet', Column A
    print("Processing 'WorkingSheet' entries in Column A...")
    for row in working_sheet.iter_rows(min_row=2, max_row=working_sheet.max_row, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value is not None:
            key = str(cell_value).strip()
            if key in perk_data:
                # Add the date from 'Perk' workbook to Column W
                working_sheet.cell(row=row[0].row, column=col_w_index, value=perk_data[key])
                print(f"Date '{perk_data[key]}' added for entry '{cell_value}' in Row {row[0].row}.")
            else:
                print(f"No match for entry '{cell_value}' in Row {row[0].row}.")

    # Save the updated workbook
    try:
        print("Saving the updated 'WorkingSheet' workbook...")
        working_wb.save(new_file_path)
        print(f"Workbook saved to {new_file_path}.")
    except Exception as e:
        raise ValueError(f"Failed to save workbook: {e}")


import sys


def main():
    if len(sys.argv) != 5:
        print("Usage: python Perk.py <source_file_path> <perk_file_path> <new_file_path> <working_sheet_name>")
        return

    source_file_path = sys.argv[1]
    perk_file_path = sys.argv[2]
    new_file_path = sys.argv[3]
    working_sheet_name = sys.argv[4]

    try:
        add_perk_dates_to_working_sheet(source_file_path, perk_file_path, new_file_path, working_sheet_name)
        print("Processing complete.")
    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    main()
