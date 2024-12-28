import openpyxl
from flask import Flask, request, render_template
import os

app = Flask(__name__)

# Load workbook and specified sheet
def load_workbook_and_sheet(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"The sheet '{sheet_name}' does not exist.")
    sheet = wb[sheet_name]
    return wb, sheet

# Get unique entries from Column G
def get_unique_entries_in_column(sheet, column_index=7):
    unique_entries = set()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
        cell_value = row[0].value
        if cell_value is not None:
            unique_entries.add(cell_value)
    return sorted(unique_entries)

# Move rows based on selected values
def move_rows_based_on_selection(sheet, selected_values, new_sheet_name="WorkingSheet"):
    wb = sheet.parent
    if new_sheet_name not in wb.sheetnames:
        new_sheet = wb.create_sheet(new_sheet_name)
        for col_idx, cell in enumerate(sheet[1], start=1):
            new_sheet.cell(row=1, column=col_idx, value=cell.value)
    else:
        new_sheet = wb[new_sheet_name]

    rows_to_delete = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[6].value in selected_values:
            next_row = new_sheet.max_row + 1
            for col_idx, cell in enumerate(row, start=1):
                new_sheet.cell(row=next_row, column=col_idx, value=cell.value)
            rows_to_delete.append(row[0].row)

    for row_num in reversed(rows_to_delete):
        sheet.delete_rows(row_num)

# Save workbook
def save_workbook(wb, file_path):
    wb.save(file_path)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # Handle file upload
        file = request.files.get("file")
        sheet_name = request.form.get("sheet_name")

        if not file or not sheet_name:
            return "Please upload a file and provide a sheet name.", 400

        # Save uploaded file
        file_path = os.path.join("uploads", file.filename)
        os.makedirs("uploads", exist_ok=True)
        file.save(file_path)

        try:
            wb, sheet = load_workbook_and_sheet(file_path, sheet_name)
            unique_entries = get_unique_entries_in_column(sheet)

            if not unique_entries:
                return "No unique entries found in Column G.", 400

            return render_template("unique_entries.html", unique_entries=unique_entries, file_path=file_path, sheet_name=sheet_name)
        except Exception as e:
            return f"Error: {e}", 500

    return render_template("index.html")

@app.route("/process_selection", methods=["POST"])
def process_selection():
    selected_values = request.form.getlist("selected_values")
    file_path = request.form.get("file_path")
    sheet_name = request.form.get("sheet_name")

    if not selected_values or not file_path or not sheet_name:
        return "Missing required inputs.", 400

    try:
        wb, sheet = load_workbook_and_sheet(file_path, sheet_name)
        move_rows_based_on_selection(sheet, selected_values)
        save_workbook(wb, file_path)

        return f"Rows with selected values moved to 'WorkingSheet'."
    except Exception as e:
        return f"Error: {e}", 500

if __name__ == "__main__":
    app.run(debug=True)
