import openpyxl
import shutil
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


def load_workbook(file_path):
    """
    Load an Excel workbook.
    """
    return openpyxl.load_workbook(file_path)


def add_perk_dates_to_working_sheet(source_file_path, perk_file_path, new_file_path, working_sheet_name="WorkingSheet"):
    """
    Add dates from 'Perk' workbook (Col J, ITS793 sheet) to entries in Column W of the WorkingSheet
    if they match entries in Column A of the WorkingSheet with Column D of Perk workbook.
    """
    print(f"Creating a copy of '{source_file_path}' as '{new_file_path}'...")
    shutil.copyfile(source_file_path, new_file_path)

    print("Loading 'WorkingSheet' workbook...")
    working_wb = load_workbook(new_file_path)
    if working_sheet_name not in working_wb.sheetnames:
        raise ValueError(f"The sheet '{working_sheet_name}' does not exist in the working workbook.")
    working_sheet = working_wb[working_sheet_name]

    print("Loading 'Perk' workbook...")
    perk_wb = load_workbook(perk_file_path)
    if "ITS793" not in perk_wb.sheetnames:
        raise ValueError("The 'Perk' workbook does not contain a sheet named 'ITS793'.")
    perk_sheet = perk_wb["ITS793"]

    # Extract all entries from Column D and corresponding dates from Column J in 'Perk' workbook into a dictionary
    print("Extracting entries from 'Perk' workbook...")
    perk_data = {}
    for row in perk_sheet.iter_rows(min_row=2, max_row=perk_sheet.max_row, min_col=4, max_col=10):
        key_value = row[0].value  # Column D (key)
        date_value = row[6].value  # Column J (date)
        if key_value is not None and date_value is not None:
            perk_data[str(key_value).strip()] = date_value

    # Add a new column header for "Perk_tag" in Column W if not already present
    print("Adding 'Perk_tag' column to 'WorkingSheet'...")
    col_w_index = 23  # Column W is the 23rd column
    if working_sheet.cell(row=1, column=col_w_index).value != "Perk_tag":
        working_sheet.cell(row=1, column=col_w_index, value="Perk_tag")

    # Iterate over each row in 'WorkingSheet', Column A
    print("Processing 'WorkingSheet' entries in Column A...")
    for row in working_sheet.iter_rows(min_row=2, max_row=working_sheet.max_row, min_col=1, max_col=1):
        cell_value = row[0].value
        if cell_value is not None:
            key = str(cell_value).strip()
            if key in perk_data:
                # Add the date from 'Perk' workbook to Column W
                working_sheet.cell(row=row[0].row, column=col_w_index, value=perk_data[key])
                print(f"Date '{perk_data[key]}' added for entry '{cell_value}' in Row {row[0].row}.")
            else:
                print(f"No match for entry '{cell_value}' in Row {row[0].row}.")

    # Save the updated workbook
    print("Saving the updated 'WorkingSheet' workbook...")
    working_wb.save(new_file_path)
    print(f"Workbook saved to {new_file_path}.")


def browse_file(entry):
    """
    Open a file dialog and insert the selected file path into the entry.
    """
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    entry.delete(0, tk.END)
    entry.insert(0, file_path)


def browse_folder(entry):
    """
    Open a folder dialog and insert the selected folder path into the entry.
    """
    folder_path = filedialog.askdirectory()
    entry.delete(0, tk.END)
    entry.insert(0, folder_path)


def start_processing(source_entry, perk_entry, dest_entry, dest_file_entry, progress_bar):
    """
    Start the Perk processing with the provided file paths.
    """
    source_file = source_entry.get()
    perk_file = perk_entry.get()
    destination_folder = dest_entry.get()
    destination_file = dest_file_entry.get()

    if not source_file or not perk_file or not destination_folder or not destination_file:
        messagebox.showerror("Error", "Please provide all required inputs.")
        return

    destination_path = os.path.join(destination_folder, destination_file)

    try:
        progress_bar.start()
        add_perk_dates_to_working_sheet(source_file, perk_file, destination_path, "WorkingSheet")
        progress_bar.stop()
        progress_bar["value"] = 100
        messagebox.showinfo("Success", f"Processing complete. File saved at: {destination_path}")
    except Exception as e:
        progress_bar.stop()
        messagebox.showerror("Error", f"An error occurred: {e}")


def main():
    """
    Main GUI for Perk script.
    """
    root = tk.Tk()
    root.title("Perk Processing Tool")

    tk.Label(root, text="Source File:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
    source_entry = tk.Entry(root, width=50)
    source_entry.grid(row=0, column=1, padx=10, pady=5)
    tk.Button(root, text="Browse", command=lambda: browse_file(source_entry)).grid(row=0, column=2, padx=10, pady=5)

    tk.Label(root, text="Perk File:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
    perk_entry = tk.Entry(root, width=50)
    perk_entry.grid(row=1, column=1, padx=10, pady=5)
    tk.Button(root, text="Browse", command=lambda: browse_file(perk_entry)).grid(row=1, column=2, padx=10, pady=5)

    tk.Label(root, text="Destination Folder:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
    dest_entry = tk.Entry(root, width=50)
    dest_entry.grid(row=2, column=1, padx=10, pady=5)
    tk.Button(root, text="Browse", command=lambda: browse_folder(dest_entry)).grid(row=2, column=2, padx=10, pady=5)

    tk.Label(root, text="Destination File Name:").grid(row=3, column=0, padx=10, pady=5, sticky="e")
    dest_file_entry = tk.Entry(root, width=50)
    dest_file_entry.grid(row=3, column=1, padx=10, pady=5)

    progress_bar = ttk.Progressbar(root, orient="horizontal", mode="determinate", length=400)
    progress_bar.grid(row=4, column=0, columnspan=3, pady=10)

    tk.Button(root, text="Start Processing", command=lambda: start_processing(
        source_entry, perk_entry, dest_entry, dest_file_entry, progress_bar
    )).grid(row=5, column=0, columnspan=3, pady=10)

    root.mainloop()


if __name__ == "__main__":
    main()
