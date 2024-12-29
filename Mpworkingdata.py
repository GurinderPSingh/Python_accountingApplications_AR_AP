import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


def load_workbook_and_sheet(file_path, sheet_name):
    """
    Load the workbook and specified sheet.
    """
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"The sheet '{sheet_name}' does not exist.")
    sheet = wb[sheet_name]
    return wb, sheet


def get_unique_entries_in_column(sheet, column_index=7):
    """
    Get unique entries from the specified column in a sheet.
    """
    unique_entries = set()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
        cell_value = row[0].value
        if cell_value is not None:
            unique_entries.add(cell_value)
    return sorted(unique_entries)


def move_rows_based_on_selection(sheet, selected_values, new_sheet_name="WorkingSheet"):
    """
    Move rows based on the selected values in Column G to a new sheet.
    """
    print(f"Processing rows for selection: {selected_values}")

    # Create or get the new sheet
    wb = sheet.parent
    if new_sheet_name not in wb.sheetnames:
        new_sheet = wb.create_sheet(new_sheet_name)
        # Copy headers from the source sheet
        for col_idx, cell in enumerate(sheet[1], start=1):
            new_sheet.cell(row=1, column=col_idx, value=cell.value)
        print(f"'{new_sheet_name}' sheet created.")
    else:
        new_sheet = wb[new_sheet_name]
        print(f"'{new_sheet_name}' sheet already exists.")

    # Identify rows to move
    rows_to_delete = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[6].value in selected_values:  # Column G is the 7th column (index 6)
            # Move this row to the new sheet
            next_row = new_sheet.max_row + 1
            for col_idx, cell in enumerate(row, start=1):
                new_sheet.cell(row=next_row, column=col_idx, value=cell.value)
            rows_to_delete.append(row[0].row)

    # Delete rows from the source sheet
    for row_num in reversed(rows_to_delete):
        sheet.delete_rows(row_num)

    print(f"Rows with values {selected_values} moved to '{new_sheet_name}' and deleted from 'Page 2'.")


def save_workbook(wb, file_path):
    """
    Save the workbook to the specified file path.
    """
    wb.save(file_path)
    print(f"Workbook saved to {file_path}.")


def process_working_data(file_path, sheet_name, selected_values, progress_bar):
    """
    Main function to process the working data from the specified file and sheet.

    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to process.
        selected_values (list): List of selected unique entries from Column G.
        progress_bar (ttk.Progressbar): Progress bar to update during the process.
    """
    try:
        # Load workbook and sheet
        wb, sheet = load_workbook_and_sheet(file_path, sheet_name)

        # Move rows based on selections
        move_rows_based_on_selection(sheet, selected_values)

        # Save the updated workbook
        save_workbook(wb, file_path)

        # Update progress bar
        progress_bar["value"] = 100
        messagebox.showinfo("Success", "Processing complete!")
    except Exception as e:
        progress_bar["value"] = 0
        messagebox.showerror("Error", f"An error occurred: {e}")
        raise


# GUI Components
def browse_file(entry):
    """
    Allow the user to browse for a file and update the entry widget.
    """
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    entry.delete(0, tk.END)
    entry.insert(0, file_path)


def start_processing(file_entry, progress_bar):
    """
    Start the processing of the working data.
    """
    file_path = file_entry.get()

    if not file_path:
        messagebox.showerror("Error", "Please select a file.")
        return

    try:
        wb, sheet = load_workbook_and_sheet(file_path, "Page 2")
        unique_entries = get_unique_entries_in_column(sheet)

        if not unique_entries:
            messagebox.showinfo("Info", "No unique entries found in Column G.")
            return

        # Create a selection dialog for unique entries
        selection_window = tk.Toplevel()
        selection_window.title("Select Unique Values")
        selection_window.geometry("400x300")

        tk.Label(selection_window, text="Select Unique Values from Column G:").pack(pady=10)

        selected_values = []

        def add_to_selected_values():
            selected_values.clear()
            selected_indices = listbox.curselection()
            selected_values.extend([unique_entries[idx] for idx in selected_indices])
            selection_window.destroy()

            # Process the data after selection
            process_working_data(file_path, "Page 2", selected_values, progress_bar)

        listbox = tk.Listbox(selection_window, selectmode="multiple")
        for entry in unique_entries:
            listbox.insert(tk.END, entry)
        listbox.pack(pady=10, fill="both", expand=True)

        tk.Button(selection_window, text="Confirm Selection", command=add_to_selected_values).pack(pady=10)

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


# Main GUI
def main():
    root = tk.Tk()
    root.title("Mpworkingdata Tool")
    root.geometry("500x300")

    tk.Label(root, text="Select the Source File:").pack(pady=10)
    file_entry = tk.Entry(root, width=50)
    file_entry.pack(pady=5)
    tk.Button(root, text="Browse", command=lambda: browse_file(file_entry)).pack(pady=5)

    progress_bar = ttk.Progressbar(root, orient="horizontal", mode="determinate", length=400)
    progress_bar.pack(pady=20)

    tk.Button(root, text="Start Processing", command=lambda: start_processing(file_entry, progress_bar)).pack(pady=20)

    root.mainloop()


if __name__ == "__main__":
    main()
