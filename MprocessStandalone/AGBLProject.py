import openpyxl
import shutil
import os
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

def load_workbook(file_path):
    """
    Load an Excel workbook.
    """
    return openpyxl.load_workbook(file_path)

def normalize(value):
    """
    Normalize cell values by stripping unnecessary characters and whitespace.
    """
    if value is None:
        return None
    return str(value).strip().lstrip('=').strip('"')

def process_agbl_project(agbl_file_path, mi_file_path, destination_path):
    """
    Process AGBL project logic, updating AGBL_MI with data from MI.
    """
    # Create a copy of the AGBL source file to avoid overwriting
    print(f"Creating a copy of '{agbl_file_path}' as '{destination_path}'...")
    shutil.copyfile(agbl_file_path, destination_path)

    print("Loading AGBL workbook...")
    agbl_wb = load_workbook(destination_path)
    agbl_sheet = agbl_wb[agbl_wb.sheetnames[0]]

    print("Loading MI workbook...")
    mi_wb = load_workbook(mi_file_path)
    mi_sheet = mi_wb[mi_wb.sheetnames[0]]

    # Add headers to AGBL_MI sheet
    agbl_sheet["J1"].value = "Person Name"
    agbl_sheet["K1"].value = "Type"
    agbl_sheet["L1"].value = "Script run date"
    agbl_sheet["M1"].value = "Count"
    agbl_sheet["N1"].value = "INVI_GL_NOS"
    agbl_sheet["O1"].value = "All Matches from Col F"

    print("Processing rows in AGBL workbook...")
    today_date = datetime.now().strftime('%Y-%m-%d')

    for row_agbl in range(2, agbl_sheet.max_row + 1):
        value_agbl_col_a = normalize(agbl_sheet[f"A{row_agbl}"].value)
        matched_values_col_f = []
        match_count = 0
        matched_value_col_m = None
        matched_value_col_g = None

        if value_agbl_col_a:
            for row_mi in range(2, mi_sheet.max_row + 1):
                value_mi_col_a = normalize(mi_sheet[f"A{row_mi}"].value)

                if value_agbl_col_a == value_mi_col_a:
                    value_col_g = normalize(mi_sheet[f"G{row_mi}"].value)
                    value_col_f = normalize(mi_sheet[f"F{row_mi}"].value)

                    if value_col_g and len(value_col_g) >= 14 and value_col_g[12] == "9" and value_col_g[13] == "9":
                        match_count += 1
                        matched_value_col_m = mi_sheet[f"M{row_mi}"].value
                        matched_value_col_g = value_col_g
                        if value_col_f:
                            matched_values_col_f.append(value_col_f)

            # Update AGBL workbook
            if match_count > 0:
                agbl_sheet[f"J{row_agbl}"].value = matched_value_col_m
                agbl_sheet[f"K{row_agbl}"].value = "Project"
                agbl_sheet[f"L{row_agbl}"].value = today_date
                agbl_sheet[f"M{row_agbl}"].value = match_count
                agbl_sheet[f"N{row_agbl}"].value = matched_value_col_g
                agbl_sheet[f"O{row_agbl}"].value = ";".join(matched_values_col_f)

    print("Saving the updated AGBL workbook...")
    agbl_wb.save(destination_path)
    print(f"Workbook saved to {destination_path}.")

# def browse_file(entry):
#     """
#     Open a file dialog and insert the selected file path into the entry.
#     """
#     file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
#     entry.delete(0, tk.END)
#     entry.insert(0, file_path)
#
# def browse_folder(entry):
#     """
#     Open a folder dialog and insert the selected folder path into the entry.
#     """
#     folder_path = filedialog.askdirectory()
#     entry.delete(0, tk.END)
#     entry.insert(0, folder_path)
#
# def start_processing(source_agbl_entry, source_mi_entry, dest_entry, dest_file_entry, progress_bar):
#     """
#     Start the AGBL processing with the provided file paths.
#     """
#     agbl_file = source_agbl_entry.get()
#     mi_file = source_mi_entry.get()
#     destination_folder = dest_entry.get()
#     destination_file = dest_file_entry.get()
#
#     if not agbl_file or not mi_file or not destination_folder or not destination_file:
#         messagebox.showerror("Error", "Please provide all required inputs.")
#         return
#
#     destination_path = os.path.join(destination_folder, destination_file)
#
#     try:
#         progress_bar.start()
#         process_agbl_project(agbl_file, mi_file, destination_path)
#         progress_bar.stop()
#         progress_bar["value"] = 100
#         messagebox.showinfo("Success", f"Processing complete. File saved at: {destination_path}")
#     except Exception as e:
#         progress_bar.stop()
#         messagebox.showerror("Error", f"An error occurred: {e}")

def main():
    """
    Main GUI for AGBLProject script.
    """
    root = tk.Tk()
    root.title("AGBL Project Processing Tool")

    # tk.Label(root, text="AGBL_MI File:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
    # source_agbl_entry = tk.Entry(root, width=50)
    # source_agbl_entry.grid(row=0, column=1, padx=10, pady=5)
    # tk.Button(root, text="Browse", command=lambda: browse_file(source_agbl_entry)).grid(row=0, column=2, padx=10, pady=5)
    #
    # tk.Label(root, text="MI File:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
    # source_mi_entry = tk.Entry(root, width=50)
    # source_mi_entry.grid(row=1, column=1, padx=10, pady=5)
    # tk.Button(root, text="Browse", command=lambda: browse_file(source_mi_entry)).grid(row=1, column=2, padx=10, pady=5)
    #
    # tk.Label(root, text="Destination Folder:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
    # dest_entry = tk.Entry(root, width=50)
    # dest_entry.grid(row=2, column=1, padx=10, pady=5)
    # tk.Button(root, text="Browse", command=lambda: browse_folder(dest_entry)).grid(row=2, column=2, padx=10, pady=5)
    #
    # tk.Label(root, text="Destination File Name:").grid(row=3, column=0, padx=10, pady=5, sticky="e")
    # dest_file_entry = tk.Entry(root, width=50)
    # dest_file_entry.grid(row=3, column=1, padx=10, pady=5)
    #
    # progress_bar = ttk.Progressbar(root, orient="horizontal", mode="determinate", length=400)
    # progress_bar.grid(row=4, column=0, columnspan=3, pady=10)
    #
    # tk.Button(root, text="Start Processing", command=lambda: start_processing(
    #     source_agbl_entry, source_mi_entry, dest_entry, dest_file_entry, progress_bar
    # )).grid(row=5, column=0, columnspan=3, pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
