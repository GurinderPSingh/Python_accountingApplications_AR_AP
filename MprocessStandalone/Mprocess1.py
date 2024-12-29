import openpyxl
import os
from datetime import datetime
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox

# Function Definitions (as previously provided)
# For brevity, assume all previously defined functions like
# load_workbook_and_sheet, copy_rows, create_page_2_sheet, etc., are implemented here.

# Example function for loading workbook and sheet
def load_workbook_and_sheet(file_path, sheet_name):
    print(f"Loading workbook from: {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
    sheet = wb[sheet_name]
    print(f"Workbook loaded. Accessing sheet: {sheet_name}.")
    return wb, sheet

# Load the workbook and the specified sheet
# def load_workbook_and_sheet(file_path, sheet_name):
#     print(f"Loading workbook from: {file_path}...")
#     wb = openpyxl.load_workbook(file_path)
#     sheet = wb[sheet_name]
#     print(f"Workbook loaded. Accessing sheet: {sheet_name}.")
#     return wb, sheet


# Copy rows below "Name" until "ITS380" is found
def copy_rows(sheet):
    print("Searching for rows between 'Name' and 'ITS380'...")
    rows_to_copy = []
    start_row = None
    is_copying = False

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            if cell.value == "Name":
                # Begin copying after "Name"
                is_copying = True
                start_row = cell.row + 1  # Start from the next row
                break

        # Collect rows if we're in copying mode
        if is_copying:
            for r in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, max_col=sheet.max_column):
                if any(cell.value == "ITS380" for cell in r):  # Stop if we hit "ITS380"
                    is_copying = False
                    break
                rows_to_copy.append([cell.value for cell in r])

    print(f"Found and copied {len(rows_to_copy)} rows.")
    return rows_to_copy


# Create or get "Page 2" sheet
def create_page_2_sheet(wb):
    print("Checking if 'Page 2' exists...")
    if "Page 2" not in wb.sheetnames:
        wb.create_sheet("Page 2")
        print("'Page 2' sheet created.")
    else:
        print("'Page 2' sheet already exists.")
    page_2_sheet = wb["Page 2"]
    return page_2_sheet


# Paste copied rows into "Page 2"
def paste_rows_to_page_2(page_2_sheet, rows_to_copy):
    print(f"Pasting {len(rows_to_copy)} rows into 'Page 2'...")
    for row_idx, row_data in enumerate(rows_to_copy, start=2):  # Start from row 2 to leave space for headers
        for col_idx, cell_value in enumerate(row_data, start=1):
            page_2_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
    print("Rows pasted successfully.")


# Find the last filled cell in column A
def find_last_filled_cell_in_col_a(sheet):
    print("Searching for the last filled cell in column A...")

    last_filled_row = None
    # Loop through rows in column A (1st column) from bottom to top
    for row_idx in range(sheet.max_row, 0, -1):  # Iterate from the last row upwards
        cell_value = sheet.cell(row=row_idx, column=1).value  # Column A is the 1st column
        if cell_value is not None:  # If the cell is not empty
            last_filled_row = row_idx
            break

    if last_filled_row:
        print(f"Last filled cell in column A is at row {last_filled_row}.")
        return last_filled_row
    else:
        print("No filled cells found in column A.")
        return None


# Function to delete the row corresponding to the last filled cell in column A
def delete_row_for_last_filled_cell_in_col_a(sheet, last_filled_row):
    if last_filled_row:
        print(f"Deleting row {last_filled_row} in 'Page 2'...")
        sheet.delete_rows(last_filled_row)
        print(f"Row {last_filled_row} deleted.")
    else:
        print("No row to delete (no filled cells in column A).")


# Delete all empty columns in the sheet
def delete_empty_columns(sheet):
    print("Deleting empty columns from 'Page 2'...")

    # Loop through columns from right to left
    last_non_empty_col = sheet.max_column
    while last_non_empty_col > 0 and all(
            sheet.cell(row=row, column=last_non_empty_col).value is None for row in range(1, sheet.max_row + 1)):
        last_non_empty_col -= 1

    # Now, delete empty columns from right to left until the last non-empty column
    for col in range(last_non_empty_col, 0, -1):
        if all(sheet.cell(row=row, column=col).value is None for row in range(1, sheet.max_row + 1)):
            sheet.delete_cols(col)
            print(f"Deleted empty column {col}")

    print("Empty columns deleted.")


# Function to check empty cells in column A and print the corresponding row numbers
def print_empty_cells_in_col_a(sheet):
    print("Checking for empty cells in column A...")

    empty_rows = []
    for row_idx in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_idx, column=1).value
        if cell_value is None:  # If the cell is empty
            empty_rows.append(row_idx)

    if empty_rows:
        print(f"Empty cells found in the following rows of column A: {empty_rows}")
    else:
        print("No empty cells found in column A.")

    return empty_rows


# Function to delete rows that correspond to empty cells in column A
def delete_rows_for_empty_cells_in_col_a(sheet, empty_rows):
    print("Deleting rows corresponding to empty cells in column A...")
    # Delete rows from bottom to top to avoid shifting issues
    for row in reversed(empty_rows):
        sheet.delete_rows(row)
        print(f"Deleted row {row}.")

    print("Rows corresponding to empty cells in column A deleted.")


# Save the workbook after changes to a new path
def save_workbook(wb, destination_folder, original_filename):
    print("Saving the modified workbook...")
    # Extract original file name and append "_modified" to it
    base_name = os.path.splitext(original_filename)[0]
    new_file_path = os.path.join(destination_folder, f"{base_name}_modified.xlsx")
    wb.save(new_file_path)
    print(f"Workbook saved successfully at {new_file_path}")
    return new_file_path


# Insert an empty row at the top (Row 1) in "Page 2" and display a message
def insert_empty_row_at_top(page_2_sheet):
    print("Inserting an empty row at the top of 'Page 2'...")
    page_2_sheet.insert_rows(1)  # Insert an empty row at the top (row 1)
    print("Empty row inserted at the top.")


# Function to assign column names to cells A1 to V1
def assign_column_names_to_header(page_2_sheet):
    column_names = [
        "Account ID", "Name", "Account Balance", "Program", "Balance By Term", "Late Fees",
        "Term", "Course Start", "Course End", "PP NEXT", "Withdrawal Date", "AWARD Start",
        "AWARD End", "CSL/MSL Start", "CSL/MSL End", "CSL", "CSL2", "EMPMB", "EMPM2", "EMPM3",
        "OTHER", "Deposit Balance"
    ]

    print("Assigning column names to cells A1 to V1...")

    for col_idx, col_name in enumerate(column_names, start=1):
        page_2_sheet.cell(row=1, column=col_idx, value=col_name)

    print("Column names assigned to A1 to V1.")


# Function to convert the date format in Column H ("Date.Month.year") to a common format (e.g., MM/DD/YYYY)
# Function to convert dates in Column H ("Day.Month.Year") to a common format (e.g., MM/DD/YYYY)
import calendar

def convert_column_h_dates(sheet):
    print("Converting dates in Column H...")

    # Mapping of month names to their numeric equivalents
    month_name_to_num = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
    }

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=8, max_col=8):  # Column H is the 8th column
        for cell in row:
            if cell.value:  # Check if the cell has a value
                try:
                    date_parts = str(cell.value).split(".")
                    if len(date_parts) == 3:
                        # Extract day, month, and year (two-digit year)
                        year = int(date_parts[0])  # Day is always numeric
                        month_part = date_parts[1]
                        day = int(date_parts[2])  # This is the 2-digit year (e.g., 24 for 2024)

                        # If the month part is a string (e.g., "Jul"), convert to a number
                        if month_part.isalpha():  # Check if the month part is a string (like "Jul")
                            month = month_name_to_num.get(month_part.capitalize())  # Get month number from the mapping
                            if month is None:
                                raise ValueError(f"Invalid month name: {month_part}")
                        else:
                            month = int(month_part)  # If it's numeric, just convert it to an integer

                        # Update year to be in the format 20yy (e.g., "24" becomes "2024")
                        year = 2000 + year  # Convert two-digit year to full year

                        # Reformat the date into "MM/dd/yyyy" format
                        formatted_date = f"{month:02}/{day:02}/{year}"
                        cell.value = formatted_date
                except Exception as e:
                    print(f"Error processing date in column H at row {cell.row}: {e}")
            else:
                continue  # If the cell is empty, leave it empty.

    print("Date conversion for Column H completed.")




# Function to convert dates in columns I to O ("Month/Date/Year") to a common format (e.g., MM/DD/YYYY)

def convert_columns_i_to_o_dates(sheet):
    print("Converting dates in Columns I to O...")

    for col in range(9, 16):  # Columns I to O are from column 9 to column 15
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    try:
                        date_str = str(cell.value).strip()

                        # Check if the date format matches the expected MonthDay/Year (e.g., "Sep26/24")
                        if "/" in date_str:
                            month_day, year_part = date_str.split("/")  # Split at "/"
                            month_part = month_day[:3]  # First 3 letters are the month (e.g., "Sep")
                            day_part = month_day[3:]  # Remaining part is the day (e.g., "26")
                            year = int("20" + year_part)  # Prefix the year with "20" (e.g., "24" becomes "2024")

                            # Mapping of month names to their numeric equivalents
                            month_name_to_num = {
                                "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
                                "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
                            }

                            # Convert month part to month number
                            month = month_name_to_num.get(month_part.capitalize())
                            if month is None:
                                raise ValueError(f"Invalid month name: {month_part}")

                            # Convert day part to integer
                            day = int(day_part)

                            # Reformat the date into MM/DD/YYYY format
                            formatted_date = datetime(year, month, day).strftime("%m/%d/%Y")
                            cell.value = formatted_date

                        else:
                            raise ValueError(f"Invalid date format: {date_str}")

                    except Exception as e:
                        print(f"Error processing date in column {cell.column_letter} at row {cell.row}: {e}")

    print("Date conversion for Columns I to O completed.")
# Function to convert dates in Column O to a common format (e.g., MM/DD/YYYY)

# Function to integrate the new date conversion logic into your main script
def convert_dates(sheet):
    print("Starting the date conversion process...")
    convert_column_h_dates(sheet)
    convert_columns_i_to_o_dates(sheet)
    print("Date conversion completed.")


#Function to sort the sheet based on column E (from smallest to largest)



def move_negative_values_to_credit_balance(sheet):
    print("Searching for negative values in Column E...")

    # Create or get "CreditBalance" sheet
    credit_balance_sheet = create_or_get_credit_balance_sheet(sheet.parent)

    # List to hold rows to delete from Page 2
    rows_to_delete = []

    # Loop through all rows starting from row 2 (to skip header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        e_value = row[4].value  # Column E (index 4)
        c_value = row[2].value  # Column C (index 2)

        if isinstance(e_value, (int, float)) and e_value < 0:  # Check if the value in column E is negative
            if e_value == c_value:  # Check if value in E matches value in C
                # Move this row to "CreditBalance"
                move_row_to_credit_balance(credit_balance_sheet, row)

                # Add this row to the list of rows to delete
                rows_to_delete.append(row[0].row)

    # Delete rows from Page 2 after moving them
    delete_rows_from_page_2(sheet, rows_to_delete)

    # Now move rows where both Column E and Column C are negative but not equal to each other
    move_negative_non_equal_values_to_cb2terms(sheet)

    # Move rows where Column E <= 10 to "LowestBalance"
    move_lowest_balance_to_new_sheet(sheet)

    print("Search complete. Negative values moved to 'CreditBalance' and deleted from 'Page 2'.")


def create_or_get_credit_balance_sheet(wb):
    print("Checking if 'CreditBalance' sheet exists...")

    # Check if the "CreditBalance" sheet exists
    if "CreditBalance" not in wb.sheetnames:
        wb.create_sheet("CreditBalance")
        print("'CreditBalance' sheet created.")

        # Copy headers from "Page 2" to "CreditBalance"
        page_2_sheet = wb["Page 2"]
        copy_headers_to_credit_balance(page_2_sheet, wb["CreditBalance"])
    else:
        print("'CreditBalance' sheet already exists.")

    # Return the "CreditBalance" sheet
    credit_balance_sheet = wb["CreditBalance"]
    return credit_balance_sheet


# Function to copy headers from "Page 2" to "CreditBalance"
def copy_headers_to_credit_balance(page_2_sheet, credit_balance_sheet):
    print("Copying headers from 'Page 2' to 'CreditBalance'...")

    # Iterate over the columns in row 1 of "Page 2" and copy them to "CreditBalance"
    for col_idx, cell in enumerate(page_2_sheet[1], start=1):  # Row 1 contains headers
        credit_balance_sheet.cell(row=1, column=col_idx, value=cell.value)

    print("Headers copied to 'CreditBalance'.")


# Function to move the row data to the "CreditBalance" sheet
def move_row_to_credit_balance(credit_balance_sheet, row):
    print(f"Moving row {row[0].row} to 'CreditBalance' sheet...")

    # Find the next available row in the "CreditBalance" sheet
    next_row = credit_balance_sheet.max_row + 1

    # Append the row data to the "CreditBalance" sheet
    for col_idx, cell in enumerate(row, start=1):
        credit_balance_sheet.cell(row=next_row, column=col_idx, value=cell.value)

    print(f"Row {row[0].row} moved to 'CreditBalance' sheet.")

# Function to create or get "LowestBalance" sheet
def create_or_get_lowest_balance_sheet(wb):
    print("Checking if 'LowestBalance' sheet exists...")

    # Check if the "LowestBalance" sheet exists
    if "LowestBalance" not in wb.sheetnames:
        wb.create_sheet("LowestBalance")
        print("'LowestBalance' sheet created.")

        # Copy headers from "Page 2" to "LowestBalance"
        page_2_sheet = wb["Page 2"]
        copy_headers_to_lowest_balance(page_2_sheet, wb["LowestBalance"])
    else:
        print("'LowestBalance' sheet already exists.")

    # Return the "LowestBalance" sheet
    lowest_balance_sheet = wb["LowestBalance"]
    return lowest_balance_sheet


# Function to copy headers from "Page 2" to "LowestBalance"
def copy_headers_to_lowest_balance(page_2_sheet, lowest_balance_sheet):
    print("Copying headers from 'Page 2' to 'LowestBalance'...")

    # Iterate over the columns in row 1 of "Page 2" and copy them to "LowestBalance"
    for col_idx, cell in enumerate(page_2_sheet[1], start=1):  # Row 1 contains headers
        lowest_balance_sheet.cell(row=1, column=col_idx, value=cell.value)

    print("Headers copied to 'LowestBalance'.")


# Function to move the row data to the "LowestBalance" sheet
def move_row_to_lowest_balance(lowest_balance_sheet, row):
    print(f"Moving row {row[0].row} to 'LowestBalance' sheet...")

    # Find the next available row in the "LowestBalance" sheet
    next_row = lowest_balance_sheet.max_row + 1

    # Append the row data to the "LowestBalance" sheet
    for col_idx, cell in enumerate(row, start=1):
        lowest_balance_sheet.cell(row=next_row, column=col_idx, value=cell.value)

    print(f"Row {row[0].row} moved to 'LowestBalance' sheet.")


# Function to move rows where Column E <= 10 to "LowestBalance" sheet
def move_lowest_balance_to_new_sheet(sheet):
    print("Searching for rows where Column E <= 10...")

    # Create or get "LowestBalance" sheet
    lowest_balance_sheet = create_or_get_lowest_balance_sheet(sheet.parent)

    # List to hold rows to delete from Page 2
    rows_to_delete = []

    # Loop through all rows starting from row 2 (to skip header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        e_value = row[4].value  # Column E (index 4)

        if isinstance(e_value, (int, float)) and 0 < e_value <= 10:  # Check if the value in column E is <= 10
            # Move this row to "LowestBalance"
            move_row_to_lowest_balance(lowest_balance_sheet, row)

            # Add this row to the list of rows to delete
            rows_to_delete.append(row[0].row)

    # Delete rows from Page 2 after moving them
    delete_rows_from_page_2(sheet, rows_to_delete)

    print("Rows with Column E <= 10 moved to 'LowestBalance' and deleted from 'Page 2'.")


# Function to move rows where both Column E and Column C are negative but not equal to "CB2Terms" sheet
def move_negative_non_equal_values_to_cb2terms(sheet):
    print("Searching for rows where Column E and Column C are negative but not equal...")

    # Create or get "CB2Terms" sheet
    cb2terms_sheet = create_or_get_cb2terms_sheet(sheet.parent)

    # List to hold rows to delete from Page 2
    rows_to_delete = []

    # Loop through all rows starting from row 2 (to skip header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        e_value = row[4].value  # Column E (index 4)
        c_value = row[2].value  # Column C (index 2)

        if isinstance(e_value, (int, float)) and e_value < 0 and isinstance(c_value, (int, float)) and c_value < 0:
            if e_value != c_value:  # Check if value in E does not equal value in C
                # Move this row to "CB2Terms"
                move_row_to_cb2terms(cb2terms_sheet, row)

                # Add this row to the list of rows to delete
                rows_to_delete.append(row[0].row)

    # Delete rows from Page 2 after moving them
    delete_rows_from_page_2(sheet, rows_to_delete)

    print("Rows with negative values in Column E and C (but not equal) moved to 'CB2Terms' and deleted from 'Page 2'.")


# Function to create or get "CB2Terms" sheet
def create_or_get_cb2terms_sheet(wb):
    print("Checking if 'CB2Terms' sheet exists...")

    # Check if the "CB2Terms" sheet exists
    if "CB2Terms" not in wb.sheetnames:
        wb.create_sheet("CB2Terms")
        print("'CB2Terms' sheet created.")

        # Copy headers from "Page 2" to "CB2Terms"
        page_2_sheet = wb["Page 2"]
        copy_headers_to_cb2terms(page_2_sheet, wb["CB2Terms"])
    else:
        print("'CB2Terms' sheet already exists.")

    # Return the "CB2Terms" sheet
    cb2terms_sheet = wb["CB2Terms"]
    return cb2terms_sheet


# Function to copy headers from "Page 2" to "CB2Terms"
def copy_headers_to_cb2terms(page_2_sheet, cb2terms_sheet):
    print("Copying headers from 'Page 2' to 'CB2Terms'...")

    # Iterate over the columns in row 1 of "Page 2" and copy them to "CB2Terms"
    for col_idx, cell in enumerate(page_2_sheet[1], start=1):  # Row 1 contains headers
        cb2terms_sheet.cell(row=1, column=col_idx, value=cell.value)

    print("Headers copied to 'CB2Terms'.")


# Function to move the row data to the "CB2Terms" sheet
def move_row_to_cb2terms(cb2terms_sheet, row):
    print(f"Moving row {row[0].row} to 'CB2Terms' sheet...")

    # Find the next available row in the "CB2Terms" sheet
    next_row = cb2terms_sheet.max_row + 1

    # Append the row data to the "CB2Terms" sheet
    for col_idx, cell in enumerate(row, start=1):
        cb2terms_sheet.cell(row=next_row, column=col_idx, value=cell.value)

    print(f"Row {row[0].row} moved to 'CB2Terms' sheet.")


# Function to delete rows from "Page 2" after they have been moved to "CreditBalance" or "CB2Terms" or "LowestBalance"
def delete_rows_from_page_2(sheet, rows_to_delete):
    # Sort rows in reverse order to avoid messing up row indices after deletion
    rows_to_delete.sort(reverse=True)

    # Delete each row in the list
    for row_num in rows_to_delete:
        sheet.delete_rows(row_num)
        print(f"Row {row_num} deleted from 'Page 2'.")

        # Function to move rows where Column J is not empty to "PPLN" sheet
def move_non_empty_column_j_to_new_sheet(sheet):
    print("Searching for rows where Column J is not empty...")

    # Create or get "PPLN" sheet
    ppln_sheet = create_or_get_ppln_sheet(sheet.parent)

    # List to hold rows to delete from Page 2
    rows_to_delete = []

    # Loop through all rows starting from row 2 (to skip header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        j_value = row[9].value  # Column J (index 9)

        if j_value is not None and j_value != "":  # Check if Column J is not empty
            # Move this row to "PPLN"
            move_row_to_ppln(ppln_sheet, row)

            # Add this row to the list of rows to delete
            rows_to_delete.append(row[0].row)

    # Delete rows from Page 2 after moving them
    delete_rows_from_page_2(sheet, rows_to_delete)

    print("Rows with Column J not empty moved to 'PPLN' and deleted from 'Page 2'.")

# Function to create or get "PPLN" sheet
def create_or_get_ppln_sheet(wb):
    print("Checking if 'PPLN' sheet exists...")
    if "PPLN" not in wb.sheetnames:
        wb.create_sheet("PPLN")
        print("'PPLN' sheet created.")
        # Copy headers from "Page 2" to "PPLN"
        page_2_sheet = wb["Page 2"]
        copy_headers_to_ppln(page_2_sheet, wb["PPLN"])
    else:
        print("'PPLN' sheet already exists.")
    ppln_sheet = wb["PPLN"]
    return ppln_sheet

# Function to copy headers from "Page 2" to "PPLN"
def copy_headers_to_ppln(page_2_sheet, ppln_sheet):
    print("Copying headers from 'Page 2' to 'PPLN'...")
    for col_idx, cell in enumerate(page_2_sheet[1], start=1):  # Row 1 contains headers
        ppln_sheet.cell(row=1, column=col_idx, value=cell.value)
    print("Headers copied to 'PPLN'.")

# Function to move the row data to the "PPLN" sheet
def move_row_to_ppln(ppln_sheet, row):
    print(f"Moving row {row[0].row} to 'PPLN' sheet...")

    # Find the next available row in the "PPLN" sheet
    next_row = ppln_sheet.max_row + 1

    # Append the row data to the "PPLN" sheet
    for col_idx, cell in enumerate(row, start=1):
        ppln_sheet.cell(row=next_row, column=col_idx, value=cell.value)

    print(f"Row {row[0].row} moved to 'PPLN' sheet.")

# Function to move rows where Column G contains "No term" to "No term" sheet
def move_no_term_to_new_sheet(sheet):
    print("Searching for rows where Column G contains 'No term'...")

    # Create or get "No term" sheet
    no_term_sheet = create_or_get_no_term_sheet(sheet.parent)

    # List to hold rows to delete from Page 2
    rows_to_delete = []

    # Loop through all rows starting from row 2 (to skip header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        g_value = row[6].value  # Column G (index 6)

        if isinstance(g_value, str) and "no term" in g_value.lower():  # Case-insensitive check for "No term"
            # Move this row to "No term"
            move_row_to_no_term(no_term_sheet, row)

            # Add this row to the list of rows to delete
            rows_to_delete.append(row[0].row)

    # Delete rows from Page 2 after moving them
    delete_rows_from_page_2(sheet, rows_to_delete)

    print("Rows with 'No term' in Column G moved to 'No term' and deleted from 'Page 2'.")

# Function to create or get "No term" sheet
def create_or_get_no_term_sheet(wb):
    print("Checking if 'No term' sheet exists...")
    if "No term" not in wb.sheetnames:
        wb.create_sheet("No term")
        print("'No term' sheet created.")
        # Copy headers from "Page 2" to "No term"
        page_2_sheet = wb["Page 2"]
        copy_headers_to_no_term(page_2_sheet, wb["No term"])
    else:
        print("'No term' sheet already exists.")
    no_term_sheet = wb["No term"]
    return no_term_sheet

# Function to copy headers from "Page 2" to "No term"
def copy_headers_to_no_term(page_2_sheet, no_term_sheet):
    print("Copying headers from 'Page 2' to 'No term'...")
    for col_idx, cell in enumerate(page_2_sheet[1], start=1):  # Row 1 contains headers
        no_term_sheet.cell(row=1, column=col_idx, value=cell.value)
    print("Headers copied to 'No term'.")

# Function to move the row data to the "No term" sheet
def move_row_to_no_term(no_term_sheet, row):
    print(f"Moving row {row[0].row} to 'No term' sheet...")

    # Find the next available row in the "No term" sheet
    next_row = no_term_sheet.max_row + 1

    # Append the row data to the "No term" sheet
    for col_idx, cell in enumerate(row, start=1):
        no_term_sheet.cell(row=next_row, column=col_idx, value=cell.value)

    print(f"Row {row[0].row} moved to 'No term' sheet.")



# Function to move rows where "(RRC)" is found in Column B to the "RRC" sheet
def move_rrc_rows_to_new_sheet(sheet):
    print("Searching for rows where Column B contains '(RRC)'...")

    # Create or get "RRC" sheet
    rrc_sheet = create_or_get_rrc_sheet(sheet.parent)

    # List to hold rows to delete from Page 2
    rows_to_delete = []

    # Loop through all rows starting from row 2 (to skip header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        b_value = row[1].value  # Column B (index 1)

        # Check if Column B contains '(RRC)' (case-insensitive)
        if b_value and isinstance(b_value, str) and "(rrc)" in b_value.lower():
            # Move this row to "RRC"
            move_row_to_rrc(rrc_sheet, row)

            # Add this row to the list of rows to delete
            rows_to_delete.append(row[0].row)

    # Delete rows from Page 2 after moving them
    delete_rows_from_page_2(sheet, rows_to_delete)

    print("Rows with '(RRC)' in Column B moved to 'RRC' and deleted from 'Page 2'.")

# Function to create or get "RRC" sheet
def create_or_get_rrc_sheet(wb):
    print("Checking if 'RRC' sheet exists...")

    # Check if the "RRC" sheet exists
    if "RRC" not in wb.sheetnames:
        wb.create_sheet("RRC")
        print("'RRC' sheet created.")

        # Copy headers from "Page 2" to "RRC"
        page_2_sheet = wb["Page 2"]
        copy_headers_to_rrc(page_2_sheet, wb["RRC"])
    else:
        print("'RRC' sheet already exists.")

    # Return the "RRC" sheet
    rrc_sheet = wb["RRC"]
    return rrc_sheet

# Function to copy headers from "Page 2" to "RRC"
def copy_headers_to_rrc(page_2_sheet, rrc_sheet):
    print("Copying headers from 'Page 2' to 'RRC'...")

    # Iterate over the columns in row 1 of "Page 2" and copy them to "RRC"
    for col_idx, cell in enumerate(page_2_sheet[1], start=1):  # Row 1 contains headers
        rrc_sheet.cell(row=1, column=col_idx, value=cell.value)

    print("Headers copied to 'RRC'.")

# Function to move the row data to the "RRC" sheet
def move_row_to_rrc(rrc_sheet, row):
    print(f"Moving row {row[0].row} to 'RRC' sheet...")

    # Find the next available row in the "RRC" sheet
    next_row = rrc_sheet.max_row + 1

    # Append the row data to the "RRC" sheet
    for col_idx, cell in enumerate(row, start=1):
        rrc_sheet.cell(row=next_row, column=col_idx, value=cell.value)

    print(f"Row {row[0].row} moved to 'RRC' sheet.")

# Function to delete rows from "Page 2" after they have been moved
# def delete_rows_from_page_2(sheet, rows_to_delete):
#     # Sort rows in reverse order to avoid messing up row indices after deletion
#     rows_to_delete.sort(reverse=True)
#
#     # Delete each row in the list
#     for row_num in rows_to_delete:
#         sheet.delete_rows(row_num)
#         print(f"Row {row_num} deleted from 'Page 2'.")


def move_rows_based_on_column_date(sheet, column_index, new_tab_name):
    """
    Move rows where the date in the specified column is greater than the current date
    to a new sheet. Deletes the rows from the source sheet after moving them.
    """
    print(f"Processing rows for column {column_index} and moving rows to '{new_tab_name}'...")

    # Create or get the new sheet
    new_sheet = create_or_get_sheet_with_headers(sheet.parent, new_tab_name, sheet)

    # List to hold rows to delete from Page 2
    rows_to_delete = []

    # Get current date
    current_date = datetime.now()

    # Loop through all rows starting from row 2 (to skip header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        cell = row[column_index - 1]  # Adjusting for 0-based index

        # Handle dates explicitly
        cell_value = cell.value
        if cell.is_date:
            # Ensure the value is a datetime object
            if isinstance(cell_value, datetime) and cell_value > current_date:
                # Move this row to the new sheet
                move_row_to_new_sheet(new_sheet, row)
                rows_to_delete.append(row[0].row)
        elif isinstance(cell_value, str):
            try:
                # Attempt to parse date strings
                parsed_date = datetime.strptime(cell_value, "%m/%d/%Y")
                if parsed_date > current_date:
                    # Move this row to the new sheet
                    move_row_to_new_sheet(new_sheet, row)
                    rows_to_delete.append(row[0].row)
            except ValueError:
                print(f"Invalid date format in cell {cell.coordinate}: {cell_value}")
                continue

    # Delete rows from Page 2 after moving them
    delete_rows_from_page_2(sheet, rows_to_delete)

    print(f"Rows with dates in column {column_index} greater than the current date moved to '{new_tab_name}'.")




def move_row_to_new_sheet(new_sheet, row):
    """
    Move a single row to the specified new sheet.
    """
    print(f"Moving row {row[0].row} to '{new_sheet.title}'...")

    # Find the next available row in the new sheet
    next_row = new_sheet.max_row + 1

    # Append the row data to the new sheet
    for col_idx, cell in enumerate(row, start=1):
        new_sheet.cell(row=next_row, column=col_idx, value=cell.value)

    print(f"Row {row[0].row} moved to '{new_sheet.title}'.")


def create_or_get_sheet_with_headers(workbook, sheet_name, source_sheet):
    """
    Creates a new sheet with the same headers as the source sheet or returns an existing sheet.
    """
    if sheet_name not in workbook.sheetnames:
        # Create a new sheet
        new_sheet = workbook.create_sheet(sheet_name)
        # Copy headers from the source sheet
        for col_idx, cell in enumerate(source_sheet[1], start=1):  # Header is assumed to be in row 1
            new_sheet.cell(row=1, column=col_idx, value=cell.value)
        print(f"'{sheet_name}' sheet created with headers.")
    else:
        # Get the existing sheet
        new_sheet = workbook[sheet_name]
        print(f"'{sheet_name}' sheet already exists.")
    return new_sheet


# def delete_rows_from_page_2(sheet, rows_to_delete):
#     """
#     Delete specified rows from the sheet.
#     """
#     # Sort rows in reverse order to avoid messing up row indices after deletion
#     rows_to_delete.sort(reverse=True)
#
#     # Delete each row in the list
#     for row_num in rows_to_delete:
#         sheet.delete_rows(row_num)
#         print(f"Row {row_num} deleted from 'Page 2'.")
def process_pending_fa(sheet):
    """
    Process columns P to U in 'Page 2', searching for '-P' and '-L', and move corresponding rows to 'PendingFA'
    and 'PendingFA-L', respectively. Delete moved rows from 'Page 2'.
    """
    print("Processing columns P to U for '-P' and '-L'...")

    # Create or get the 'PendingFA' and 'PendingFA-L' sheets
    pending_fa_sheet = create_or_get_sheet_with_headers(sheet.parent, "PendingFA", sheet)
    pending_fa_l_sheet = create_or_get_sheet_with_headers(sheet.parent, "PendingFA-L", sheet)

    # Columns P to U correspond to indices 16 to 21
    for col_index in range(16, 22):
        print(f"Searching column {openpyxl.utils.get_column_letter(col_index)} for '-P' and '-L'...")

        # Lists to hold rows to delete from Page 2
        rows_to_delete = []

        # Loop through all rows starting from row 2 (to skip header)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
            cell = row[0]  # Single cell in this column
            if cell.value and isinstance(cell.value, str):
                if "-P" in cell.value:
                    # Move this row to 'PendingFA'
                    move_row_to_new_sheet(pending_fa_sheet, sheet[cell.row])
                    rows_to_delete.append(cell.row)
                elif "-L" in cell.value:
                    # Move this row to 'PendingFA-L'
                    move_row_to_new_sheet(pending_fa_l_sheet, sheet[cell.row])
                    rows_to_delete.append(cell.row)

        # Delete rows from Page 2 after processing the column
        delete_rows_from_page_2(sheet, rows_to_delete)
        print(f"Completed processing column {openpyxl.utils.get_column_letter(col_index)}.")

    print("All relevant rows moved to 'PendingFA' and 'PendingFA-L' and deleted from 'Page 2'.")



from openpyxl.styles import PatternFill


def highlight_rows_with_p_in_tabs(workbook, exclude_tabs=None):
    """
    Highlight rows in columns P to U with '-P' in all tabs except the specified exclude_tabs.

    Args:
        workbook (Workbook): The workbook containing the sheets.
        exclude_tabs (list): List of sheet names to exclude from processing. Defaults to ['PendingFA', 'Page 1', 'Page 2'].
    """
    # Default to common excluded tabs if none provided
    if exclude_tabs is None:
        exclude_tabs = ["PendingFA", "Page 1", "Page 2"]

    print(f"Highlighting rows with '-P' in columns P to U, excluding tabs: {exclude_tabs}")

    # Define the orange fill for highlighting
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Loop through all sheets in the workbook, excluding specified tabs
    for sheet_name in workbook.sheetnames:
        if sheet_name in exclude_tabs:
            continue  # Skip the excluded sheets

        sheet = workbook[sheet_name]
        print(f"Processing sheet: {sheet_name}")

        # Loop through columns P to U (16 to 21)
        for col_index in range(16, 22):
            # Loop through all rows starting from row 2 (to skip header)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
                cell = row[0]  # Single cell in this column
                if cell.value and isinstance(cell.value, str) and "-P" in cell.value:
                    # Highlight the entire row
                    row_idx = cell.row
                    for col in range(1, sheet.max_column + 1):  # Highlight all columns in the row
                        sheet.cell(row=row_idx, column=col).fill = orange_fill
                    print(f"Row {row_idx} in sheet '{sheet_name}' highlighted.")

    print("Highlighting complete.")


def move_non_zero_in_col_v(sheet):
    """
    Search for non-zero values in column V in 'Page 2' and move corresponding rows to 'DepositBalance'.
    If the 'DepositBalance' sheet does not exist, it is created with the same headers as 'Page 2'.

    Args:
        sheet (Worksheet): The source worksheet ('Page 2').
    """
    print("Searching for non-zero values in column V and moving rows to 'DepositBalance'...")

    # Create or get the 'DepositBalance' sheet
    deposit_balance_sheet = create_or_get_sheet_with_headers(sheet.parent, "DepositBalance", sheet)

    # List to hold rows to delete from Page 2
    rows_to_delete = []

    # Loop through all rows starting from row 2 (to skip header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=22, max_col=22):  # Column V is the 22nd column
        cell = row[0]  # Single cell in this column
        if cell.value and isinstance(cell.value, (int, float)) and cell.value != 0:
            # Move this row to 'DepositBalance'
            move_row_to_new_sheet(deposit_balance_sheet, sheet[cell.row])

            # Add this row to the list of rows to delete
            rows_to_delete.append(cell.row)

    # Delete rows from Page 2 after processing
    delete_rows_from_page_2(sheet, rows_to_delete)

    print("Rows with non-zero values in column V moved to 'DepositBalance' and deleted from 'Page 2'.")


def move_rows_with_col_i_past_date(sheet):
    """
    Search for rows in column I with dates less than the current date in 'Page 2',
    move corresponding rows to the 'termEndDate' sheet (creating it with the same headers if needed),
    and delete those rows from 'Page 2'.

    Args:
        sheet (Worksheet): The source worksheet ('Page 2').
    """
    print("Processing column I for dates less than the current date and moving rows to 'termEndDate'...")

    # Create or get the 'termEndDate' sheet
    term_end_date_sheet = create_or_get_sheet_with_headers(sheet.parent, "termEndDate", sheet)

    # List to hold rows to delete from Page 2
    rows_to_delete = []

    # Get the current date
    current_date = datetime.now()

    # Loop through all rows starting from row 2 (to skip header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=9, max_col=9):  # Column I is the 9th column
        cell = row[0]  # Single cell in this column
        cell_value = cell.value

        if cell.is_date:
            # Ensure the value is a datetime object and compare
            if isinstance(cell_value, datetime) and cell_value < current_date:
                # Move this row to 'termEndDate'
                move_row_to_new_sheet(term_end_date_sheet, sheet[cell.row])
                rows_to_delete.append(cell.row)
        elif isinstance(cell_value, str):
            try:
                # Attempt to parse date strings
                parsed_date = datetime.strptime(cell_value, "%m/%d/%Y")
                if parsed_date < current_date:
                    # Move this row to 'termEndDate'
                    move_row_to_new_sheet(term_end_date_sheet, sheet[cell.row])
                    rows_to_delete.append(cell.row)
            except ValueError:
                # Handle invalid date formats
                print(f"Invalid date format in cell {cell.coordinate}: {cell_value}")
                continue

    # Delete rows from Page 2 after processing
    delete_rows_from_page_2(sheet, rows_to_delete)

    print("Rows with column I dates less than the current date moved to 'termEndDate' and deleted from 'Page 2'.")

import openpyxl

def move_empty_col_h_rows(sheet, new_sheet_name="noCourseStartDate"):
    """
    Find rows with empty cells in Column H in 'Page 2',
    move them to a new sheet named 'noCourseStartDate',
    and delete them from 'Page 2'.

    Args:
        sheet (Worksheet): The source worksheet ('Page 2').
        new_sheet_name (str): The name of the new sheet to create or get.
    """
    print("Processing rows with empty cells in Column H...")

    # Create or get the 'noCourseStartDate' sheet
    no_course_start_date_sheet = create_or_get_sheet_with_headers(sheet.parent, new_sheet_name, sheet)

    # List to hold rows to delete from Page 2
    rows_to_delete = []

    # Loop through all rows starting from row 2 (to skip header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=8, max_col=8):  # Column H is the 8th column
        cell = row[0]  # Single cell in this column
        if cell.value is None or cell.value == "":
            # Move this row to 'noCourseStartDate'
            move_row_to_new_sheet(no_course_start_date_sheet, sheet[cell.row])
            rows_to_delete.append(cell.row)

    # Delete rows from Page 2 after processing
    rows_to_delete.sort(reverse=True)  # Sort in reverse order to avoid shifting rows during deletion
    for row_num in rows_to_delete:
        sheet.delete_rows(row_num)
        print(f"Deleted row {row_num} from 'Page 2'.")

    print(f"Rows with empty Column H moved to '{new_sheet_name}' and deleted from 'Page 2'.  This operations is performed at the end after all sorting and before creating workingsheet of your selection")



# Main function to execute the steps


def process_source_file(file_path, sheet_name, destination_folder):
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"File '{file_path}' does not exist.")

    if not os.path.isdir(destination_folder):
        raise FileNotFoundError(f"Directory '{destination_folder}' does not exist.")

    try:
        print("Starting processing...")
        wb, sheet = load_workbook_and_sheet(file_path, sheet_name)

        rows_to_copy = copy_rows(sheet)
        if not rows_to_copy:
            print("No rows found to copy.")
            return

        page_2_sheet = create_page_2_sheet(wb)
        paste_rows_to_page_2(page_2_sheet, rows_to_copy)

        last_filled_row_in_col_a = find_last_filled_cell_in_col_a(page_2_sheet)
        delete_row_for_last_filled_cell_in_col_a(page_2_sheet, last_filled_row_in_col_a)

        delete_empty_columns(page_2_sheet)

        empty_rows = print_empty_cells_in_col_a(page_2_sheet)
        delete_rows_for_empty_cells_in_col_a(page_2_sheet, empty_rows)

        insert_empty_row_at_top(page_2_sheet)
        assign_column_names_to_header(page_2_sheet)
        convert_dates(page_2_sheet)

        move_negative_values_to_credit_balance(page_2_sheet)
        move_rrc_rows_to_new_sheet(page_2_sheet)
        move_non_empty_column_j_to_new_sheet(page_2_sheet)
        move_no_term_to_new_sheet(page_2_sheet)
        move_rows_based_on_column_date(page_2_sheet, column_index=8, new_tab_name="FutureTerm")
        move_rows_based_on_column_date(page_2_sheet, column_index=15, new_tab_name="CSLPerk")
        process_pending_fa(page_2_sheet)
        highlight_rows_with_p_in_tabs(wb)
        move_non_zero_in_col_v(page_2_sheet)
        move_rows_with_col_i_past_date(page_2_sheet)
        move_empty_col_h_rows(page_2_sheet)

        save_workbook(wb, destination_folder, os.path.basename(file_path))
        print("Processing complete.")
        messagebox.showinfo("Success", "Processing complete!")
    except Exception as e:
        print(f"Error during processing: {e}")
        messagebox.showerror("Error", f"Error during processing: {e}")

# GUI Implementation
def browse_file(entry):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    entry.delete(0, tk.END)
    entry.insert(0, file_path)

def browse_folder(entry):
    folder_path = filedialog.askdirectory()
    entry.delete(0, tk.END)
    entry.insert(0, folder_path)

def start_processing(file_entry, sheet_entry, folder_entry):
    file_path = file_entry.get()
    sheet_name = sheet_entry.get()
    destination_folder = folder_entry.get()

    if not file_path or not sheet_name or not destination_folder:
        messagebox.showerror("Error", "Please fill in all fields.")
        return

    try:
        process_source_file(file_path, sheet_name, destination_folder)
    except Exception as e:
        print(f"Error: {e}")
        messagebox.showerror("Error", f"An error occurred: {e}")

# Main GUI Window
def main():
    root = tk.Tk()
    root.title("Excel Processing Tool")

    tk.Label(root, text="Source File:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
    file_entry = tk.Entry(root, width=50)
    file_entry.grid(row=0, column=1, padx=10, pady=5)
    tk.Button(root, text="Browse", command=lambda: browse_file(file_entry)).grid(row=0, column=2, padx=10, pady=5)

    tk.Label(root, text="Sheet Name:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
    sheet_entry = tk.Entry(root, width=50)
    sheet_entry.grid(row=1, column=1, padx=10, pady=5)

    tk.Label(root, text="Destination Folder:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
    folder_entry = tk.Entry(root, width=50)
    folder_entry.grid(row=2, column=1, padx=10, pady=5)
    tk.Button(root, text="Browse", command=lambda: browse_folder(folder_entry)).grid(row=2, column=2, padx=10, pady=5)

    tk.Button(root, text="Start Processing", command=lambda: start_processing(file_entry, sheet_entry, folder_entry)).grid(
        row=3, column=0, columnspan=3, pady=10
    )

    root.mainloop()

if __name__ == "__main__":
    main()
